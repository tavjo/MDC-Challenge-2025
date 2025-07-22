import os
import logging
import numpy as np
import faiss  # New import for FAISS
from typing import List, Optional
from pathlib import Path
import PyPDF2
import openpyxl
from docx import Document
from io import BytesIO
import pypandoc

from monitoring import log_all_methods
from models import DocumentInformation, Interpretation, DocumentText
from services.document_service import DocumentService
from services.azure_embedding_client import AzureEmbeddingClient

# Configure logging
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
ENABLE_TIMING = os.environ.get("ENABLE_TIMING", "true").lower() in ("true", "1", "yes")


@log_all_methods(enabled=ENABLE_TIMING)
class DocumentProcessor:
    """
    A class to process documents and return the top relevant chunks.
    """

    def __init__(
        self,
        chunk_size: int = 500,
        top_n: int = 3,
        embedding_batch_size: int = 64,
        embedding_client: Optional[AzureEmbeddingClient] = None,
        document_service: Optional[DocumentService] = None,
    ):
        """
        Initialize the DocumentProcessor.
        """

        # Use the injected AzureEmbeddingClient
        self.embedding_client = embedding_client
        self.document_service = document_service
        self.chunk_size = int(chunk_size)
        self.top_n = int(top_n)
        self.embedding_batch_size = int(embedding_batch_size)
        self.skipped_documents = []
        self.file_type_handlers = {
            ".pdf": self._pdf_to_text,
            ".docx": self._docx_to_text,
            ".xlsx": self._xlsx_to_text,
            ".doc": self._doc_to_text,
        }

    def process(self, interpretation: Interpretation, document_information: DocumentInformation) -> DocumentText:
        """
        Process the document and return the top relevant chunks.
        """

        logging.info(f"Processing document: {document_information.document_name}.")
        try:
            raw_text = self.retrieve_relevant_document_text(document_url=document_information.document_url)
            chunks = self.split_into_chunks(raw_text=raw_text)
            target_document_text = interpretation.refined_query.hypothetical_response_value

            # Use FAISS to retrieve the top relevant chunks based on cosine similarity
            top_chunks = self.get_top_relevant_chunks(chunks=chunks, target_document_text=target_document_text)

            processed_document = DocumentText(
                document_information=document_information,
                relevant_passages=top_chunks,
            )
            return processed_document
        except Exception as e:
            logging.error(f"Error processing {document_information.document_name}: {e}")
            self.skipped_documents.append({
                "document_name": document_information.document_name,
                "error": str(e),
            })
            raise

    def retrieve_relevant_document_text(self, document_url: str) -> Optional[str]:
        """
        Retrieve and process the document text by delegating to the
        retrieve_document and return_document_text methods.
        """

        logging.info("Retrieving document text from %s.", document_url)
        # Retrieve the raw document content from the document service
        file_content = self.retrieve_document(document_url)
        if not file_content:
            logging.warning("No file content retrieved from %s.", document_url)
            return None

        # Process the document text using the appropriate file handler
        return self.return_document_text(document_url, file_content)

    def retrieve_document(self, document_url: str) -> Optional[bytes]:
        """
        Retrieve the raw document content from the document service.

        Parameters:
            document_url (str): The URL of the document.

        Returns:
            Optional[bytes]: The raw document content or None if retrieval fails.
        """

        if not document_url:
            logging.warning("No document URL provided.")
            return None

        logging.info("Retrieving document from %s.", document_url)
        try:
            file_content = self.document_service.retrieve_document(document_url)
            if not file_content:
                logging.warning("No file content retrieved from %s.", document_url)
            return file_content
        except Exception as e:
            logging.error("Error retrieving document from %s: %s", document_url, e)
            return None

    def return_document_text(self, document_url: str, file_content: bytes) -> Optional[str]:
        """
        Process and return the document text using the appropriate file handler.

        Parameters:
            document_url (str): The URL of the document (used to determine file type).
            file_content (bytes): The raw content of the document.

        Returns:
            Optional[str]: The processed document text or None if conversion fails.
        """

        if not file_content:
            logging.warning("No file content provided for %s.", document_url)
            return None

        # Determine file extension and select the corresponding handler.
        file_extension = Path(document_url).suffix.lower()
        handler = self.file_type_handlers.get(file_extension)
        if handler:
            try:
                return handler(file_content)
            except Exception as e:
                logging.error("Error processing document text for %s: %s", document_url, e)
                return None
        else:
            logging.warning("Unsupported file type: %s", file_extension)
            return None

    def split_into_chunks(self, raw_text: str, chunk_size: Optional[int] = None) -> List[str]:
        """
        Split the raw text into chunks of the specified size.

        Returns an empty list if the input is empty or only whitespace.
        Raises a ValueError if chunk_size is not a positive integer.
        """

        if chunk_size is None:
            chunk_size = self.chunk_size

        logging.debug("Splitting raw text into chunks of size %d.", chunk_size)

        # Check for empty or whitespace-only text.
        if not raw_text.strip():
            logging.warning("Empty text received for chunking. Returning empty list.")
            return []

        # Ensure chunk_size is a positive integer.
        if not isinstance(chunk_size, int) or chunk_size <= 0:
            logging.error("chunk_size must be a positive integer. Got: %r", chunk_size)
            raise ValueError("chunk_size must be a positive integer")

        # Compute the length once.
        text_length = len(raw_text)

        # Build the list of chunks using slicing.
        chunks = [raw_text[i: i + chunk_size] for i in range(0, text_length, chunk_size)]
        logging.debug("Created %d chunks.", len(chunks))
        return chunks

    def get_top_relevant_chunks(self, chunks: List[str], target_document_text: str) -> List[str]:
        """
        Calculate the top relevant chunks using FAISS for cosine similarity.

        Parameters:
            chunks (List[str]): List of text chunks to search among.
            target_document_text (str): The document text to use as a query.

        Returns:
            List[str]: The top N relevant chunks based on cosine similarity.

        Raises:
            ValueError: If the input embeddings are empty.
            RuntimeError: If an error occurs during embedding.
        """

        logging.info("Calculating top relevant chunks using FAISS for cosine similarity.")

        # Validate inputs
        if not chunks:
            logging.warning("No chunks provided for embedding.")
            return []
        if not target_document_text:
            logging.warning("No target document text provided for embedding.")
            return []

        # Embed the target document (query)
        try:
            query_embedding_list = self.embedding_client.embed(target_document_text)
        except Exception as e:
            logging.error(f"Failed to embed target document: {e}")
            raise RuntimeError(f"Failed to embed target document: {e}")

        if not query_embedding_list or len(query_embedding_list) == 0:
            raise ValueError("Query embedding is empty.")

        # Convert query embedding to NumPy array as float32
        query_vector = np.array(query_embedding_list[0], dtype='float32')
        query_vector = np.expand_dims(query_vector, axis=0)

        # Embed all chunks and convert to NumPy array as float32
        try:
            chunk_embedding_list = self.embedding_client.embed(chunks)
        except Exception as e:
            logging.error(f"Failed to embed chunks: {e}")
            raise RuntimeError(f"Failed to embed chunks: {e}")

        # Convert chunk_embedding_list to NumPy array as float32
        chunk_embeddings = np.array(chunk_embedding_list, dtype='float32')
        if chunk_embeddings.size == 0:
            raise ValueError("Chunk embeddings are empty.")
        logging.debug(f"Embedded {len(chunks)} chunks.")

        # Normalize embeddings (in-place normalization for efficiency)
        faiss.normalize_L2(chunk_embeddings)
        faiss.normalize_L2(query_vector)

        # Optionally adjust self.top_n if it's greater than available chunks
        top_n = min(self.top_n, len(chunks))

        # Build a FAISS index using inner product (which now equals cosine similarity)
        dimension = chunk_embeddings.shape[1]
        index = faiss.IndexFlatIP(dimension)
        index.add(chunk_embeddings)

        # Search the index for the top_n nearest neighbors
        distances, indices = index.search(query_vector, top_n)

        # Retrieve the corresponding chunks
        top_chunks = [chunks[i] for i in indices[0]]
        logging.debug("Retrieved top relevant chunks successfully.")

        return top_chunks

    def _doc_to_text(self, file_content) -> str:
        """
        Convert a DOC file to text using pypandoc.
        """

        logging.info("Converting DOC to text using pypandoc.")
        try:
            with BytesIO(file_content) as temp_file:
                temp_file.seek(0)
                output = pypandoc.convert_text(temp_file.read().decode("utf-8"), to="plain", format="doc")
                if not output.strip():
                    raise ValueError("DOC to text conversion resulted in empty output.")
                logging.debug("DOC to text conversion successful.")
                return output.strip()
        except Exception as e:
            logging.error(f"Error converting DOC to text: {e}")
            return ""

    def _pdf_to_text(self, file_content) -> str:
        """
        Convert a PDF file to text.
        """

        logging.info("Converting PDF to text.")
        try:
            # Convert file_content to bytes if it's not already bytes
            pdf_reader = PyPDF2.PdfReader(BytesIO(file_content))
            # Extract text from each page
            text = "".join(page.extract_text() for page in pdf_reader.pages)
            return text.strip()
        except Exception as e:
            logging.error(f"Error converting PDF to text: {e}")
            return ""

    def docx_to_text_pypandoc(self, file_content) -> str:
        """
        Convert a DOCX file to text using pypandoc.
        """

        logging.info("Converting DOCX to text using pypandoc.")
        try:
            # Convert file_content to bytes if it's not already bytes
            text = pypandoc.convert_text(file_content, to="plain", format="docx").strip()
            return text if text else ""
        except Exception as e:
            logging.error(f"Error converting DOCX to text with pypandoc: {e}")
            return ""

    def docx_to_text_python_docx(self, file_content) -> str:
        """
        Convert a DOCX file to text using python-docx.
        """

        logging.info("Converting DOCX to text using python-docx.")
        try:
            # Convert file_content to bytes if it's not already bytes
            document = Document(BytesIO(file_content))
            # Extract text from each paragraph
            text = "\n".join(paragraph.text for paragraph in document.paragraphs)
            return text.strip()
        except Exception as e:
            logging.error(f"Error converting DOCX to text with python-docx: {e}")
            return ""

    def _docx_to_text(self, file_content) -> str:
        """
        Tries pypandoc first, then falls back to python-docx.
        """

        # Try pypandoc first
        logging.info("Trying pypandoc first.")
        text = self.docx_to_text_pypandoc(file_content)
        if not text:
            # Fallback to python-docx if pypandoc fails or returns nothing
            logging.info("Pypandoc failed. Falling back to python-docx.")
            text = self.docx_to_text_python_docx(file_content)
        return text

    def _xlsx_to_text(self, file_content) -> str:
        """
        Convert an XLSX file to text.
        """

        # Convert file_content to bytes if it's not already bytes
        logging.info("Converting XLSX to text.")
        try:
            # Convert file_content to bytes if it's not already bytes
            workbook = openpyxl.load_workbook(filename=BytesIO(file_content), data_only=True)
            text = ""
            # Extract text from each sheet
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                text += f"\nSheet: {sheet}\n"
                # Extract text from each row
                for row in worksheet.iter_rows(values_only=True):
                    # Extract text from each cell
                    text += " ".join([str(cell) if cell else "" for cell in row]) + "\n"
            return text.strip()
        except Exception as e:
            logging.error(f"Error converting XLSX to text: {e}")
            return ""

    def _fallback_conversion(self, file_content, file_extension) -> str:
        """
        Fallback conversion for unsupported file types.
        """

        logging.info(f"Attempting fallback conversion for {file_extension}.")
        try:
            # Convert file_content to bytes if it's not already bytes
            with BytesIO(file_content) as temp_file:
                temp_file.seek(0)
                # Convert the file content to plain text
                output = pypandoc.convert_text(
                    temp_file.read().decode("utf-8"),
                    to="plain",
                    format=file_extension[1:],
                )
                return output.strip()
        except Exception as e:
            logging.error(f"Fallback conversion failed: {e}")
            return ""

    def report_skipped_documents(self):
        """
        Report the skipped documents.
        """

        # Report the skipped documents
        if self.skipped_documents:
            logging.info("Skipped Documents Report:")
            for skipped in self.skipped_documents:
                logging.info(f"Document: {skipped['document_name']}, Error: {skipped['error']}")
        else:
            logging.info("No documents were skipped.")

    def retrieve_relevant_passages_from_file(
        self,
        file_content: bytes,
        query: str,
        top_n: Optional[int] = None,
        chunk_size: Optional[int] = None,
        filename: Optional[str] = None
    ) -> List[str]:
        """
        Process the given file content (of any allowed type) along with a query string,
        and return the top k most relevant text chunks.

        Parameters:
            file_content (bytes): The raw content of the file.
            query (str): The query string to search for relevant text.
            k (Optional[int]): The number of top relevant chunks to return.
                            Defaults to self.top_n if not provided.
            filename (Optional[str]): The filename (used to determine file type).
                                    If provided, it is used to select the appropriate conversion handler.

        Returns:
            List[str]: The top k relevant text chunks.
        """
        # Use provided k or default to the instance's top_n.
        if top_n is None:
            top_n = self.top_n

        if chunk_size is None:
            chunk_size = self.chunk_size

        # Extract text using the appropriate handler based on the filename.
        if filename:
            text = self.return_document_text(document_url=filename, file_content=file_content)
        else:
            # Fallback conversion if filename is not provided.
            logging.warning("No filename provided; using fallback conversion for .txt.")
            text = self._fallback_conversion(file_content, ".txt")

        if not text:
            logging.warning("No text could be extracted from the file.")
            return []

        # Split the extracted text into manageable chunks.
        chunks = self.split_into_chunks(text, chunk_size)
        if not chunks:
            logging.warning("No chunks were created from the extracted text.")
            return []

        # Embed the query string.
        try:
            query_embedding_list = self.embedding_client.embed(query)
        except Exception as e:
            logging.error("Failed to embed the query: %s", e)
            raise RuntimeError(f"Failed to embed query: {e}")
        if not query_embedding_list or len(query_embedding_list) == 0:
            raise ValueError("Query embedding is empty.")
        query_vector = np.array(query_embedding_list[0], dtype='float32')
        query_vector = np.expand_dims(query_vector, axis=0)

        # Embed all text chunks.
        try:
            chunk_embedding_list = self.embedding_client.embed(chunks)
        except Exception as e:
            logging.error("Failed to embed chunks: %s", e)
            raise RuntimeError(f"Failed to embed chunks: {e}")
        chunk_embeddings = np.array(chunk_embedding_list, dtype='float32')
        if chunk_embeddings.size == 0:
            raise ValueError("Chunk embeddings are empty.")

        # Normalize embeddings so that inner product search equals cosine similarity.
        faiss.normalize_L2(chunk_embeddings)
        faiss.normalize_L2(query_vector)

        # Build a FAISS index using the dimension of the embeddings.
        dimension = chunk_embeddings.shape[1]
        index = faiss.IndexFlatIP(dimension)
        index.add(chunk_embeddings)

        # Ensure we request no more results than available chunks.
        top_k = min(top_n, len(chunks))
        distances, indices = index.search(query_vector, top_k)

        # Retrieve and return the corresponding chunks.
        top_chunks = [chunks[i] for i in indices[0]]
        return top_chunks
